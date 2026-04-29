from __future__ import annotations
import argparse
import os
import re
import msoffcrypto

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# --- MODELOS ---
class Severity(str, Enum):
    HIGH = "Alta"
    MEDIUM = "Media"
    LOW = "Baja"

@dataclass(frozen=True)
class Finding:
    file_path: str
    file_type: str
    risk_type: str
    detail: str
    severity: Severity

# --- CONFIGURACIÓN DE ESCANEO ---
DEFAULT_IGNORED_DIRS = {"bin", "obj", ".git", "node_modules", "$RECYCLE.BIN"}
SUPPORTED_SUFFIXES = {
    ".config", ".xml", ".json", ".env", ".ini", ".txt", ".log", ".sql",
    ".py", ".js", ".ts", ".cs", ".vb", ".docx", ".xlsx", ".pdf", ".bat", ".ps1"
}

# --- REGEX DE DETECCIÓN (OLFATO COMPLETO) ---
INTERNAL_IP_RE = re.compile(r"\b(10\.(?:\d{1,3}\.){2}\d{1,3}|192\.168\.(?:\d{1,3}\.)\d{1,3}|172\.(?:1[6-9]|2\d|3[01])\.(?:\d{1,3}\.)\d{1,3})\b")
CONNSTRING_RE = re.compile(r"(?i)\b(Server|Data Source|User ID|Password|PWD|UID)\s*=\s*[^;]+")
JWT_RE = re.compile(r"\beyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\b")
BEARER_RE = re.compile(r"(?i)\bBearer\s+([A-Za-z0-9\-_\.=]{20,})\b")

# Captura la CLAVE y el VALOR real
KEYVALUE_ASSIGN_RE = re.compile(r"(?ix)\b(?P<key>PASSWORD|PWD|SECRET|TOKEN|APIKEY|API_KEY|CONTRASENA|CONTRASEÑA|LOGIN|URL)\b\s*[:=]\s*(?P<val>['\"]?[^\s#;]{5,100}['\"]?)")

# --- MOTORES DE EXTRACCIÓN (TEXTO Y BINARIOS) ---
def extract_text(path: Path, max_bytes: int) -> str:
    suffix = path.suffix.lower()
    try:
        # Texto plano y código
        if suffix in {".config", ".xml", ".json", ".env", ".ini", ".txt", ".log", ".sql", ".py", ".js", ".ts", ".cs", ".vb", ".bat", ".ps1"}:
            data = path.read_bytes()[:max_bytes]
            for enc in ("utf-8", "latin-1", "cp1252"):
                try: return data.decode(enc)
                except: continue
            return data.decode("utf-8", errors="ignore")

        # Microsoft Word
        if suffix == ".docx":
            from docx import Document
            doc = Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)

        # Microsoft Excel
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
            return "\n".join(" ".join(str(c) for c in r if c) for s in wb.worksheets for r in s.iter_rows(values_only=True))

        # PDF
        if suffix == ".pdf":
            import pdfplumber
            with pdfplumber.open(str(path)) as pdf:
                return "\n".join(page.extract_text() or "" for page in pdf.pages)
    except:
        return ""
    return ""

def detect_content_risks(path: Path, text: str) -> list[Finding]:
    if not text: return []
    findings = []
    suffix = path.suffix.lower()
    
    # 1. Búsqueda de Contraseñas/Secretos con valor real
    for match in KEYVALUE_ASSIGN_RE.finditer(text):
        key = match.group('key')
        val = match.group('val')
        findings.append(Finding(str(path), suffix, "Secreto Expuesto", f"{key} = {val}", Severity.HIGH))
    
    # 2. Cadenas de Conexión
    if CONNSTRING_RE.search(text):
        findings.append(Finding(str(path), suffix, "Conexión DB", "Cadena SQL con credenciales", Severity.HIGH))
    
    # 3. Tokens JWT/Bearer
    if JWT_RE.search(text) or BEARER_RE.search(text):
        findings.append(Finding(str(path), suffix, "Token Acceso", "JWT/Bearer detectado", Severity.HIGH))
    
    # 4. IPs Internas
    if INTERNAL_IP_RE.search(text):
        findings.append(Finding(str(path), suffix, "IP Interna", "Infraestructura privada", Severity.LOW))
        
    return findings

# --- LÓGICA DE ESCANEO Y EXCEL ---
def scan(root: Path, excludes: set[str], max_bytes: int) -> list[Finding]:
    findings = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in excludes]
        for fn in filenames:
            p = Path(dirpath) / fn
            if p.suffix.lower() in SUPPORTED_SUFFIXES:
                content = extract_text(p, max_bytes)
                findings.extend(detect_content_risks(p, content))
    return findings

def write_xlsx(findings: list[Finding], out_path: Path):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(["Archivo", "Ext", "Tipo de Riesgo", "Detalle (Valor Encontrado)", "Severidad"])
    
    fills = {Severity.HIGH: PatternFill("solid", "FF3B30"), Severity.MEDIUM: PatternFill("solid", "FF9500"), Severity.LOW: PatternFill("solid", "007AFF")}
    
    for f in findings:
        ws.append([f.file_path, f.file_type, f.risk_type, f.detail, f.severity.value])
        for cell in ws[ws.max_row]: cell.fill = fills.get(f.severity)
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 85)
        
    wb.save(str(out_path))

def encrypt_xlsx(input_file: Path, output_file: Path, password: str):
  
    from io import BytesIO

    with open(input_file, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)

        # Crear archivo cifrado en memoria
        out = BytesIO()
        office_file.encrypt(password=password, outfile=out)

    with open(output_file, "wb") as f:
        f.write(out.getvalue())

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--password", required=True)
    args = parser.parse_args()

    print("🚀 Escaneando...")
    res = scan(Path(args.path).resolve(), DEFAULT_IGNORED_DIRS, 2000000)

    temp_file = Path(args.out).with_suffix(".temp.xlsx")

    write_xlsx(res, temp_file)

    print("🔐 Cifrando archivo...")
    encrypt_xlsx(temp_file, Path(args.out), args.password)

    temp_file.unlink()  # borrar temporal
    print(f"✅ Reporte cifrado generado en: {args.out}")
if __name__ == "__main__":
    main()
